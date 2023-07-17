# coding: utf-8
"""
Base para desarrollo de modulos externos.
Para obtener el modulo/Funcion que se esta llamando:
     GetParams("module")

Para obtener las variables enviadas desde formulario/comando Rocketbot:
    var = GetParams(variable)
    Las "variable" se define en forms del archivo package.json

Para modificar la variable de Rocketbot:
    SetVar(Variable_Rocketbot, "dato")

Para obtener una variable de Rocketbot:
    var = GetVar(Variable_Rocketbot)

Para obtener la Opcion seleccionada:
    opcion = GetParams("option")


Para instalar librerias se debe ingresar por terminal a la carpeta "libs"
    
   sudo pip install <package> -t .

"""


from pymsgbox import *
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

# This lines is to linter
# -----------------------------------
GetParams = GetParams #type:ignore
tmp_global_obj = tmp_global_obj #type:ignore
PrintException = PrintException #type:ignore
SetVar = SetVar #type:ignore
GetGlobals = GetGlobals #type:ignore
# Add modules libraries to Rocektbot
# -----------------------------------
module = GetParams("module")

if module == "Config":
    
    try:
        path_excel = GetParams('path_excel')
        list_hojas = GetParams('list_hojas')
        config = GetParams('config')
        
        list_hojas = eval(list_hojas)
        #config = {}
        wb = load_workbook(path_excel, read_only=True)

        
        config_dict = {}

        
        for sheet in list_hojas:
        
            ws = wb[sheet]

        
            for row in ws.iter_rows(min_row=2, values_only=True):  
                if row[0] is not None and row[1] is not None:  
                    config_dict[row[0]] = row[1]
                    
        import json
        config_json = json.dumps(config_dict)  
        if config_dict:
            SetVar(config, config_json)    
        wb.close()
        del wb
        import gc
        gc.collect()
        
    except Exception as e:
        PrintException()
        raise e

if module == "valor":
    key_ = GetParams("key")
    dict_ = GetParams("config")
    res = GetParams("var_")
    try:
        import json
        new_dict = json.loads(dict_) 
        valor = new_dict[key_]          
        SetVar(res, valor)
    except Exception as e:
        print("\x1B[" + "31;40mError\x1B[" + "0m")
        PrintException()
        raise e
    

if module == "Change_date_american":
    date_from = GetParams("date_from")
    res = GetParams("date_to")

    try:
        fecha_datetime = datetime.strptime(date_from, '%d-%m-%Y')
        fecha_final = fecha_datetime.strftime('%Y-%m-%d')
        if fecha_final:
            print(fecha_final)
            SetVar(res, fecha_final)
            
    except Exception as e:
        PrintException()
        raise e

if module == "delete_row_value":
    path = GetParams("path")
    value = GetParams("value")
    Column = GetParams("Column")
    try:
        df = pd.read_excel(path)
        df = df[df[Column] != value]
        df[Column] = df[Column].astype(str).str.strip()
        df.to_excel(path, index=False)
    except Exception as e:
        PrintException()
        raise e
    
if module == "transform_dates":
    path_to_excel = GetParams("path")
    date_column = GetParams("Column_name")
    original_format = GetParams("original_format")
    new_format = GetParams("new_format")
    
    try:
        # Leemos el archivo Excel
        df = pd.read_excel(path_to_excel)
    except FileNotFoundError:
        print(f"No se encontró el archivo en {path_to_excel}")
        PrintException()
        
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        PrintException()
        raise e
        
    try:
        df[date_column] = pd.to_datetime(df[date_column], format=original_format)        
        df[date_column] = df[date_column].dt.strftime(new_format)
    except KeyError:
        print(f"No se encontró la columna {date_column}")
        PrintException()
        
    except Exception as e:
        print(f"Error al transformar las fechas: {e}")
        PrintException()
        raise e
        

    try:
        df.to_excel(path_to_excel, index=False)
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")
        PrintException()
        raise e
    
if module == "keep_row_value":
    path = GetParams("path")
    hoja_from = GetParams("hoja_from")
    hoja_to  = GetParams("hoja_to")
    value = GetParams("value")
    column = GetParams("Column")
    try:
        # Leer el archivo Excel y las hojas específicas
        xls = pd.ExcelFile(path)
        if hoja_from not in xls.sheet_names or hoja_to not in xls.sheet_names:
            raise ValueError("Las hojas especificadas no existen en el archivo Excel")

        df_from = pd.read_excel(xls, hoja_from)
        df_to = pd.read_excel(xls, hoja_to)

        # Verificar que la columna exista
        if column not in df_from.columns:
            raise ValueError("La columna especificada no existe en la hoja de origen")

        # Filtrar los datos
        df_filtrado = df_from[df_from[column] == value]

        # Copiar los datos filtrados a la hoja_to
        df_to = pd.concat([df_to, df_filtrado], ignore_index=True)

        # Escribir el DataFrame resultante de nuevo al archivo Excel
        with pd.ExcelWriter(path) as writer:  
            df_to.to_excel(writer, sheet_name=hoja_to, index=False)
            df_from.to_excel(writer, sheet_name=hoja_from, index=False)

    except FileNotFoundError:
        print(f"El archivo {path} no existe")
    except PermissionError:
        print(f"No se tienen permisos de lectura/escritura para el archivo {path}")
    except Exception as e:
        print(f"Ha ocurrido un error inesperado: {str(e)}")
